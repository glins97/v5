import os
import os.path
import pickle
import codecs 
import io
import re

from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request

import openpyxl 
import pandas
import numpy as np
import os 
from copy import copy, deepcopy
import datetime

import matplotlib
import matplotlib.pyplot as plt
matplotlib.use('Agg')

from io import BytesIO
from .models import TPS, TPSAnswer, Question, QuestionAnswer

def get_row(cell):
    return int(re.search(r'(\d+?)$', cell).group(1))

def get_tps_answers_dataframe(tps):
    columns = ["xA", "xC", "xD",]
    tps_answers = TPSAnswer.objects.filter(tps=tps,)
    data = [ (answer.submission_date, answer.grade, answer.name) for answer in tps_answers ]
    return pandas.DataFrame(data, columns=columns)

def separate_students(tps):
    df = get_tps_answers_dataframe(tps).sort_values(by=(['xC', 'xA']), ascending=[False, True])
    result = {
        'SCORE_Z': pandas.DataFrame(columns=list(df.columns.values)),
        'TBL': pandas.DataFrame(columns=list(df.columns.values)),
        'CBT': pandas.DataFrame(columns=list(df.columns.values)),
    }
    if tps.campus == 'GOI':
        threshold = df.shape[0] * 0.8
        for index, row in df.iterrows():
            if index < threshold:
                result['SCORE_Z'] = result['SCORE_Z'].append(row, ignore_index=True)
            else:
                result['TBL'] = result['TBL'].append(row, ignore_index=True)
    elif tps.campus == 'BSB':
        count_score_z = 0
        count_tbl = 0
        for _, row in df.iterrows():
            if row['xC'] >= tps.max_questions * 0.8 and count_score_z < 10:
                result['SCORE_Z'] = result['SCORE_Z'].append(row, ignore_index=True)
                count_score_z += 1
            elif row['xC'] >= tps.max_questions * 0.6 and count_tbl < 20:
                result['TBL'] = result['TBL'].append(row, ignore_index=True)
                count_tbl += 1
            else:
                result['CBT'] = result['CBT'].append(row, ignore_index=True)
    elif tps.campus == 'JUA':
        count = 0
        for _, row in df.iterrows():
            if count < 20:
                result['SCORE_Z'] = result['SCORE_Z'].append(row, ignore_index=True)
                count += 1
            else:
                result['TBL'] = result['TBL'].append(row, ignore_index=True)

    return result

def calculate_statistics(df):
    statistics = {}
    statistics['MIN'] = '-'
    statistics['MAX'] = '-'
    statistics['MEAN'] = '-'
    statistics['STD'] = '-'
    if df.shape[0]:
        statistics['MIN'] = np.min(df['xC'])
        statistics['MAX'] = np.max(df['xC'])
        statistics['MEAN'] = np.mean(df['xC'])
        statistics['STD'] = np.std(df['xC'])
    return statistics

def duplicate(ws, origin, destination):
    ws[destination].font = copy(ws[origin].font)
    ws[destination].border = copy(ws[origin].border)
    ws[destination].fill = copy(ws[origin].fill)
    ws[destination].number_format = copy(ws[origin].number_format)
    ws[destination].protection = copy(ws[origin].protection)
    ws[destination].alignment = copy(ws[origin].alignment)
    if type(ws[destination]).__name__ != 'MergedCell':
        ws[destination].value = copy(ws[origin].value)

    if (ws.row_dimensions[get_row(origin)].height == 7.5):
        ws.row_dimensions[get_row(destination)].height = 7.5
    
def generate_score_z(tps):
    students = separate_students(tps)['SCORE_Z']
    print(students)
    stats = calculate_statistics(students)

    wb = openpyxl.load_workbook(filename='tps/inputs/TEMPLATE_SCORE_Z.xlsx')
    ws = wb.active
    ws['D3'] = str(tps)
    ws['G3'] = datetime.datetime.now().strftime('%d/%m/%Y')
    ws['C6'] = 'MAIOR: 0'
    ws['D6'] = 'MENOR: 0'
    ws['E6'] = 'MÉDIA: 0'
    ws['F6'] = 'DESVIO: 0'
    students = list(students.iterrows())
    if students:
        ws['C6'] = 'MAIOR: {:.2f}'.format(stats['MAX'])
        ws['D6'] = 'MENOR: {:.2f}'.format(stats['MIN'])
        ws['E6'] = 'MÉDIA: {:.2f}'.format(stats['MEAN'])
        ws['F6'] = 'DESVIO: {:.2f}'.format(stats['STD'])
        ws['C10'] = students[0][1]['xD'].upper()
        ws['G10'] = students[0][1]['xC']
        ws['B10'] = '1.  '
        ws['B10'].alignment = openpyxl.styles.Alignment(horizontal='right', vertical='center')
        ws['B10'].border = copy(ws['B5'].border)
        ws['C10'].border = copy(ws['C5'].border)
        ws['D10'].border = copy(ws['D5'].border)
        ws['E10'].border = copy(ws['E5'].border)
        ws['F10'].border = copy(ws['F5'].border)
        ws['G10'].border = copy(ws['G5'].border)
        for index, (_, student) in enumerate(students[1:]):
            row = str(11 + index)
            duplicate(ws, 'C10', 'C' + row)
            duplicate(ws, 'G10', 'G' + row)
            ws['C' + row].border = copy(ws['A1'].border)
            ws['B' + row].border = copy(ws['B5'].border)
            ws['G' + row].border = copy(ws['G5'].border)
            ws['C' + row] = student['xD'].upper()
            ws['G' + row] = student['xC']
            ws['B' + row] = '{}.  '.format(index + 2)
            ws['B' + row].alignment = openpyxl.styles.Alignment(horizontal='right', vertical='center')
            if type(ws['C' + row]).__name__ != 'MergedCell':
                ws.merge_cells('C{}:F{}'.format(row, row))
        ws['B' + str(9 + len(students))].border = copy(ws['B6'].border)
        ws['C' + str(9 + len(students))].border = copy(ws['C6'].border)
        ws['D' + str(9 + len(students))].border = copy(ws['D6'].border)
        ws['E' + str(9 + len(students))].border = copy(ws['E6'].border)
        ws['F' + str(9 + len(students))].border = copy(ws['F6'].border)
        ws['G' + str(9 + len(students))].border = copy(ws['G6'].border)

    wb.save('tps/outputs/xlsxs/SCORE_Z_' + str(tps).replace(' ', '_').upper() + '.xlsx')
    return 'tps/outputs/xlsxs/SCORE_Z_' + str(tps).replace(' ', '_').upper() + '.xlsx'

def generate_tbl(tps):
    students = separate_students(tps)['TBL']
    stats = calculate_statistics(students)

    wb = openpyxl.load_workbook(filename='tps/inputs/TEMPLATE_TBL.xlsx')
    ws = wb.active
    ws['D3'] = str(tps)
    ws['G3'] = datetime.datetime.now().strftime('%d/%m/%Y')
    ws['C6'] = 'MAIOR: 0'
    ws['D6'] = 'MENOR: 0'
    ws['E6'] = 'MÉDIA: 0'
    ws['F6'] = 'DESVIO: 0'
        
    students = list(students.iterrows())
    if students:
        ws['C6'] = 'MAIOR: {:.2f}'.format(stats['MAX'])
        ws['D6'] = 'MENOR: {:.2f}'.format(stats['MIN'])
        ws['E6'] = 'MÉDIA: {:.2f}'.format(stats['MEAN'])
        ws['F6'] = 'DESVIO: {:.2f}'.format(stats['STD'])
        ws['C10'] = students[0][1]['xD'].upper()
        ws['G10'] = students[0][1]['xC']
        ws['B10'] = '1.  '
        ws['B10'].alignment = openpyxl.styles.Alignment(horizontal='right', vertical='center')
        ws['B10'].border = copy(ws['B5'].border)
        ws['C10'].border = copy(ws['C5'].border)
        ws['D10'].border = copy(ws['D5'].border)
        ws['E10'].border = copy(ws['E5'].border)
        ws['F10'].border = copy(ws['F5'].border)
        ws['G10'].border = copy(ws['G5'].border)
        for index, (_, student) in enumerate(students[1:]):
            row = str(11 + index)
            duplicate(ws, 'C10', 'C' + row)
            duplicate(ws, 'G10', 'G' + row)
            ws['C' + row].border = copy(ws['A1'].border)
            ws['B' + row].border = copy(ws['B5'].border)
            ws['G' + row].border = copy(ws['G5'].border)
            ws['C' + row] = student['xD'].upper()
            ws['G' + row] = student['xC']
            ws['B' + row] = '{}.  '.format(index + 2)
            ws['B' + row].alignment = openpyxl.styles.Alignment(horizontal='right', vertical='center')
            if type(ws['C' + row]).__name__ != 'MergedCell':
                ws.merge_cells('C{}:F{}'.format(row, row))
        ws['B' + str(9 + len(students))].border = copy(ws['B6'].border)
        ws['C' + str(9 + len(students))].border = copy(ws['C6'].border)
        ws['D' + str(9 + len(students))].border = copy(ws['D6'].border)
        ws['E' + str(9 + len(students))].border = copy(ws['E6'].border)
        ws['F' + str(9 + len(students))].border = copy(ws['F6'].border)
        ws['G' + str(9 + len(students))].border = copy(ws['G6'].border)

    wb.save('tps/outputs/xlsxs/TBL_' + str(tps).replace(' ', '_').upper() + '.xlsx')
    return 'tps/outputs/xlsxs/TBL_' + str(tps).replace(' ', '_').upper() + '.xlsx'

def generate_cbt(tps):
    students = separate_students(tps)['CBT']
    stats = calculate_statistics(students)

    wb = openpyxl.load_workbook(filename='tps/inputs/TEMPLATE_CBT.xlsx')
    ws = wb.active
    ws['D3'] = str(tps)
    ws['G3'] = datetime.datetime.now().strftime('%d/%m/%Y')
    ws['C6'] = 'MAIOR: 0'
    ws['D6'] = 'MENOR: 0'
    ws['E6'] = 'MÉDIA: 0'
    ws['F6'] = 'DESVIO: 0'
    students = list(students.iterrows())
    if students:
        ws['C6'] = 'MAIOR: {:.2f}'.format(stats['MAX'])
        ws['D6'] = 'MENOR: {:.2f}'.format(stats['MIN'])
        ws['E6'] = 'MÉDIA: {:.2f}'.format(stats['MEAN'])
        ws['F6'] = 'DESVIO: {:.2f}'.format(stats['STD'])
        ws['C10'] = students[0][1]['xD'].upper()
        ws['G10'] = students[0][1]['xC']
        ws['B10'] = '1.  '
        ws['B10'].alignment = openpyxl.styles.Alignment(horizontal='right', vertical='center')
        ws['B10'].border = copy(ws['B5'].border)
        ws['C10'].border = copy(ws['C5'].border)
        ws['D10'].border = copy(ws['D5'].border)
        ws['E10'].border = copy(ws['E5'].border)
        ws['F10'].border = copy(ws['F5'].border)
        ws['G10'].border = copy(ws['G5'].border)
        for index, (_, student) in enumerate(students[1:]):
            row = str(11 + index)
            duplicate(ws, 'C10', 'C' + row)
            duplicate(ws, 'G10', 'G' + row)
            ws['C' + row].border = copy(ws['A1'].border)
            ws['B' + row].border = copy(ws['B5'].border)
            ws['G' + row].border = copy(ws['G5'].border)
            ws['C' + row] = student['xD'].upper()
            ws['G' + row] = student['xC']
            ws['B' + row] = '{}.  '.format(index + 2)
            ws['B' + row].alignment = openpyxl.styles.Alignment(horizontal='right', vertical='center')
            if type(ws['C' + row]).__name__ != 'MergedCell':
                ws.merge_cells('C{}:F{}'.format(row, row))
        ws['B' + str(9 + len(students))].border = copy(ws['B6'].border)
        ws['C' + str(9 + len(students))].border = copy(ws['C6'].border)
        ws['D' + str(9 + len(students))].border = copy(ws['D6'].border)
        ws['E' + str(9 + len(students))].border = copy(ws['E6'].border)
        ws['F' + str(9 + len(students))].border = copy(ws['F6'].border)
        ws['G' + str(9 + len(students))].border = copy(ws['G6'].border)

    wb.save('tps/outputs/xlsxs/CBT_' + str(tps).replace(' ', '_').upper() + '.xlsx')
    return 'tps/outputs/xlsxs/CBT_' + str(tps).replace(' ', '_').upper() + '.xlsx'

def generate_distrator(tps):
    questions = sorted(Question.objects.filter(tps=tps,), key=lambda question: question.number)
    question_answers = {q: QuestionAnswer.objects.filter(question=q) for q in questions}
    grades = [tps_answer.grade for tps_answer in TPSAnswer.objects.filter(tps=tps)]

    wb = openpyxl.load_workbook(filename='tps/inputs/TEMPLATE_DISTRATOR.xlsx')
    ws = wb.active
    ws['C3'] = str(tps)
    ws['G3'] = datetime.datetime.now().strftime('%d/%m/%Y')
    if grades:
        ws['C6'] = 'MAIOR: {:.2f}'.format(np.max(grades))
        ws['D6'] = 'MENOR: {:.2f}'.format(np.min(grades))
        ws['E6'] = 'MÉDIA: {:.2f}'.format(np.mean(grades))
        ws['F6'] = 'DESVIO: {:.2f}'.format(np.std(grades))
    else:
        ws['C6'] = 'MAIOR: -'
        ws['D6'] = 'MENOR: -'
        ws['E6'] = 'MÉDIA: -'
        ws['F6'] = 'DESVIO: -'

    row = 11
    ws[f'B{row}'] = 'Q. 1'
    for index, letter in enumerate('ABCDE'):
        column = 'CDEFG'[index]
        text = str(sum([1 for question_answer in question_answers[questions[0]] if question_answer.answer == letter]))
        if letter == questions[0].correct_answer:
            text = '   ' + text + ' ✓'
        ws[f'{column}{row}'] = text 

    for index, question in enumerate(questions[1:]):
        duplicate(ws, f'B{row + index}', f'B{row + index + 1}')
        ws[f'B{row + index + 1}'] = f'Q. {index + 2}'
        for l_index, letter in enumerate('ABCDE'):
            column = 'CDEFG'[l_index]
            duplicate(ws, f'{column}{row + index}', f'{column}{row + index + 1}')
            text = str(sum([1 for question_answer in question_answers[question] if question_answer.answer == letter]))
            if letter == question.correct_answer:
                text = '   ' + text + ' ✓'
            ws[f'{column}{row + index + 1}'] = text 

    plt.hist(grades, bins=tps.max_questions)  # `density=False` would make counts
    plt.ylabel('Alunos')
    plt.xlabel('Nota')
    
    buffer = BytesIO()
    plt.savefig(buffer, format = 'png')
    plt.clf()
    img = openpyxl.drawing.image.Image(buffer)
    img.anchor = f'B{12 + tps.max_questions}'
    ws.add_image(img)

    wb.save('tps/outputs/xlsxs/DISTRATOR_' + str(tps).replace(' ', '_').upper() + '.xlsx')
    return 'tps/outputs/xlsxs/DISTRATOR_' + str(tps).replace(' ', '_').upper() + '.xlsx'